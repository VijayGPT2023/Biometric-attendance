"""Reports: printable views, PDF and Excel export."""
import os
import json
import io
from datetime import datetime
from flask import render_template, request, current_app, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from app.extensions import db
from app.models.attendance import UploadSession
from app.models.justification import Justification
from app.models.audit import AuditLog
from app.blueprints.reports import reports_bp
from app.blueprints.attendance.serializers import deserialize_results
from app.blueprints.attendance.routes import _get_holidays_for_range, _get_holiday_names
from app.utils.helpers import group_by_department


def _load_session_data(session_uuid):
    data_path = os.path.join(current_app.config['DATA_FOLDER'], f"{session_uuid}.json")
    if not os.path.exists(data_path):
        return None, None, None, None
    with open(data_path) as f:
        data = json.load(f)
    return deserialize_results(data)


@reports_bp.route('/<session_uuid>')
@login_required
def view_report(session_uuid):
    result = _load_session_data(session_uuid)
    if not result[0]:
        flash('Session not found.')
        return redirect(url_for('auth.dashboard'))

    results, start_date, end_date, params = result
    upload_session = UploadSession.query.filter_by(session_uuid=session_uuid).first()
    office_id = upload_session.office_id if upload_session else None
    holidays = sorted(_get_holidays_for_range(start_date, end_date, office_id))
    holiday_names = _get_holiday_names(start_date, end_date, office_id)
    dept_groups = group_by_department(results)

    return render_template('reports/report.html',
                           dept_groups=dept_groups, results=results,
                           start_date=start_date, end_date=end_date,
                           params=params, holidays=holidays,
                           holiday_names=holiday_names,
                           session_id=session_uuid, stage='initial')


@reports_bp.route('/<session_uuid>/excel')
@login_required
def export_excel(session_uuid):
    """Export attendance report as Excel file."""
    result = _load_session_data(session_uuid)
    if not result[0]:
        flash('Session not found.')
        return redirect(url_for('auth.dashboard'))

    results, start_date, end_date, params = result
    dept_groups = group_by_department(results)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = 'Summary'
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='E8EAF6')
    danger_font = Font(color='C62828', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # Title
    ws.merge_cells('A1:P1')
    ws['A1'] = f'Attendance Report: {start_date.strftime("%d-%b-%Y")} to {end_date.strftime("%d-%b-%Y")}'
    ws['A1'].font = Font(bold=True, size=14)

    row = 3
    for dept_name, dept_results in dept_groups.items():
        ws.cell(row=row, column=1, value=dept_name).font = Font(bold=True, size=12)
        row += 1

        headers = ['Code', 'Name', 'Designation', 'W.Days', 'Present', 'Absent',
                   'Late', 'Early', 'Short Hrs', 'Miss Dep', 'Miss Arr',
                   'Anomalies', 'Effective', 'Leave Ded.']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for r in dept_results:
            vals = [r['emp_code'], r['emp_name'], r['designation'],
                    r['working_days'], r['present'], r['absent'],
                    r['late_arrival_count'], r['early_departure_count'],
                    r['short_hours_count'], r['missing_departure_count'],
                    r['missing_arrival_count'], r['total_anomaly_dates_raw'],
                    r['effective_anomaly_count'], r['leave_deduction']]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                if col == 14 and v > 0:
                    cell.font = danger_font
            row += 1
        row += 1

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Detail sheets per department
    for dept_name, dept_results in dept_groups.items():
        safe_name = dept_name[:30].replace('/', '-')
        ws_dept = wb.create_sheet(title=safe_name)
        ws_dept.cell(row=1, column=1, value=f'{dept_name} - Anomaly Details').font = Font(bold=True, size=12)

        row = 3
        for r in dept_results:
            if not r.get('anomaly_details'):
                continue
            ws_dept.cell(row=row, column=1, value=f"{r['emp_name']} ({r['emp_code']})").font = Font(bold=True)
            row += 1
            for col, h in enumerate(['Date', 'Day', 'Arrival', 'Departure', 'Hours', 'Anomaly Type'], 1):
                cell = ws_dept.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            for d in r['anomaly_details']:
                ws_dept.cell(row=row, column=1, value=d['date'].strftime('%d-%b-%Y'))
                ws_dept.cell(row=row, column=2, value=d['date'].strftime('%a'))
                ws_dept.cell(row=row, column=3, value=d.get('arrival', '-'))
                ws_dept.cell(row=row, column=4, value=d.get('departure', '-'))
                ws_dept.cell(row=row, column=5, value=d.get('hours', '-'))
                ws_dept.cell(row=row, column=6, value=', '.join(d.get('types', [])))
                row += 1
            row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    AuditLog.log('export_report', user_id=current_user.id,
                 resource_type='upload_session', resource_id=session_uuid,
                 details='excel')

    filename = f'Attendance_Report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/<session_uuid>/pdf')
@login_required
def export_pdf(session_uuid):
    """Export attendance report as PDF."""
    result = _load_session_data(session_uuid)
    if not result[0]:
        flash('Session not found.')
        return redirect(url_for('auth.dashboard'))

    results, start_date, end_date, params = result
    dept_groups = group_by_department(results)
    upload_session = UploadSession.query.filter_by(session_uuid=session_uuid).first()
    office_id = upload_session.office_id if upload_session else None
    holidays = sorted(_get_holidays_for_range(start_date, end_date, office_id))
    holiday_names = _get_holiday_names(start_date, end_date, office_id)

    html = render_template('reports/pdf_report.html',
                           dept_groups=dept_groups, results=results,
                           start_date=start_date, end_date=end_date,
                           params=params, holidays=holidays,
                           holiday_names=holiday_names)

    try:
        from xhtml2pdf import pisa
        output = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=output)
        output.seek(0)

        AuditLog.log('export_report', user_id=current_user.id,
                     resource_type='upload_session', resource_id=session_uuid,
                     details='pdf')

        filename = f'Attendance_Report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.pdf'
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/pdf')
    except ImportError:
        flash('PDF export requires xhtml2pdf package. Please install it.')
        return redirect(url_for('reports.view_report', session_uuid=session_uuid))
